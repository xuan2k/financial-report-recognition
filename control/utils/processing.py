import numpy as np  
import cv2
import random
import math
from PIL import Image, ImageDraw, ImageFont
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from bs4 import BeautifulSoup
import json
from scipy.spatial import KDTree
from box import *
from functools import reduce

def vis_img(image, bbox):
    bbox = np.asarray(bbox)
    bbox = bbox.astype(np.int64)
    # color = (0,0,255)
    color = (random.randint(0,255),random.randint(0,255),random.randint(0,255))
    thickness = 2
    # bbox = [200,100,400,500]
    cv2.rectangle(image, (bbox[0], bbox[1]), (bbox[2], bbox[3]), color, thickness)
    return image

def crop_img(image, bbox):
    bbox = np.asarray(bbox)
    bbox = bbox.astype(np.int64)
    # print(image.shape)
    image = image[bbox[1]:bbox[3], bbox[0]:bbox[2]]
    # print(image.shape)
    return image

def create_font(txt, sz, font_path="./doc/fonts/simfang.ttf"):
    font_size = int(sz[1] * 0.99)
    font = ImageFont.truetype(font_path, font_size, encoding="utf-8")
    length = font.getsize(txt)[0]
    if length > sz[0]:
        font_size = int(font_size * sz[0] / length)
        font = ImageFont.truetype(font_path, font_size, encoding="utf-8")
    return font

def draw_box_txt_fine(img_size, box, txt, font_path="./doc/fonts/simfang.ttf"):
    box_height = int(
        math.sqrt((box[0][0] - box[3][0])**2 + (box[0][1] - box[3][1])**2))
    box_width = int(
        math.sqrt((box[0][0] - box[1][0])**2 + (box[0][1] - box[1][1])**2))

    if box_height > 2 * box_width and box_height > 30:
        img_text = Image.new('RGB', (box_height, box_width), (255, 255, 255))
        draw_text = ImageDraw.Draw(img_text)
        if txt:
            font = create_font(txt, (box_height, box_width), font_path)
            draw_text.text([0, 0], txt, fill=(0, 0, 0), font=font)
        img_text = img_text.transpose(Image.ROTATE_270)
    else:
        img_text = Image.new('RGB', (box_width, box_height), (255, 255, 255))
        draw_text = ImageDraw.Draw(img_text)
        if txt:
            font = create_font(txt, (box_width, box_height), font_path)
            draw_text.text([0, 0], txt, fill=(0, 0, 0), font=font)

    pts1 = np.float32(
        [[0, 0], [box_width, 0], [box_width, box_height], [0, box_height]])
    pts2 = np.array(box, dtype=np.float32)
    M = cv2.getPerspectiveTransform(pts1, pts2)

    img_text = np.array(img_text, dtype=np.uint8)
    img_right_text = cv2.warpPerspective(
        img_text,
        M,
        img_size,
        flags=cv2.INTER_NEAREST,
        borderMode=cv2.BORDER_CONSTANT,
        borderValue=(255, 255, 255))
    return img_right_text

def draw_ocr_box_txt(image,
                     boxes,
                     txts=None,
                     scores=None,
                     drop_score=0.5,
                     font_path="./doc/fonts/simfang.ttf"):
    h, w = image.height, image.width
    img_left = image.copy()
    img_right = np.ones((h, w, 3), dtype=np.uint8) * 255
    random.seed(0)

    draw_left = ImageDraw.Draw(img_left)
    if txts is None or len(txts) != len(boxes):
        txts = [None] * len(boxes)
    
    for idx, (box, txt) in enumerate(zip(boxes, txts)):
        if scores is not None and scores[idx] < drop_score:
            continue
        color = (random.randint(0, 255), random.randint(0, 255),
                 random.randint(0, 255))
        box = [tuple(coor) for coor in box]
        # print(box)
        # print(color)
        draw_left.polygon(box, fill=color)
        img_right_text = draw_box_txt_fine((w, h), box, txt, font_path)
        pts = np.array(box, np.int32).reshape((-1, 1, 2))
        cv2.polylines(img_right_text, [pts], True, color, 1)
        img_right = cv2.bitwise_and(img_right, img_right_text)
    img_left = Image.blend(image, img_left, 0.5)
    img_show = Image.new('RGB', (w * 2, h), (255, 255, 255))
    img_show.paste(img_left, (0, 0, w, h))
    img_show.paste(Image.fromarray(img_right), (w, 0, w * 2, h))
    return np.array(img_show)

def binarize(image, thresh = 0.7):
    thresh *= 255
    image = image.point( lambda p: 255 if p > thresh else 0 )
    return image

# Intersect over Itself
def IoI(boxA, boxB):
    xA, yA, xB, yB = boxA
    xC, yC, xD, yD = boxB

    x_left = max(xA, xC)
    y_top = max(yA, yC)
    x_right = min(xB, xD)
    y_bottom = min(yB, yD)

    width = max(0, x_right - x_left + 1)
    height = max(0, y_bottom - y_top + 1)

    area_A = (xB - xA + 1) * (yB - yA + 1)
    area_intersection = width * height

    percentage_intersection = (area_intersection / area_A) * 100

    return percentage_intersection

def to_excel():

    excel_file = '/home/xuan/Project/OCR/sample/result/pred/demo.xlsx'
    wb = Workbook()
    ws = wb.active

    # Raw HTML table code with rowspan and colspan attributes
    # html_table = '''
    # <html><body><table><thead><tr><td>1</td><td>1</td><td rowspan=\"2\">1</td><td rowspan=\"2\">1</td><td>1</td></tr></thead><tbody><tr><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td rowspan=\"3\">1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td rowspan=\"2\">1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr></tbody></table></body></html>

    # '''

    # html_table = '''
    # <html><body><table><thead><tr><td>1</td><td rowspan=\"3\">1</td><td rowspan=\"3\">1</td><td colspan=\"2\">1</td><td colspan=\"2\">1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr></thead><tbody><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td rowspan=\"2\">1</td><td rowspan=\"2\">1</td><td colspan=\"3\">1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td rowspan=\"2\">1</td><td>1</td><td>1</td><td rowspan=\"2\">1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td rowspan=\"2\">1</td><td>1</td><td rowspan=\"2\">1</td><td rowspan=\"2\">1</td><td rowspan=\"2\">1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td><td>1</td></tr><tr><td>1</td><td>1</td><td>1</td><td colspan=\"4\">1</td></tr></tbody></table></body></html>
    # '''

    f = open("/home/xuan/Project/OCR/sample/result/pred/res_1_TSR.txt", 'r')

    data = json.load(f)

    rec_f = open("/home/xuan/Project/OCR/sample/result/text/1/res.json", "r")


    html_table = data["html"]
    bboxes = data["content_ann"]["bboxes"]
    ratio = data["ratio"]
    tmp = bboxes
    bboxes = []
    for box in tmp:
        box = np.asarray(box).astype(np.float64)
        box *= ratio
        box = list(box.astype(np.int64))
        bboxes.append(box)
 
    rec = json.load(rec_f)
    rec_boxes = rec["bboxes"]
    # Parse the HTML table using BeautifulSoup
    soup = BeautifulSoup(html_table, 'html.parser')

    # Find all the rows in the table
    rows = soup.find_all('tr')
    # print(rows)

    # Initialize an empty list to store the table data
    table_data = []
    list_a = []

    # Loop through each row in the table
    cellss = []
    count = 0
    for row in rows:
        # Find all the cells in the row
        cells = row.find_all(['td', 'th'])
        cellss.append(cells)
        # print(cells)
        
        # Extract the content of each cell
        row_data = [(bboxes[count + i]) for i in range(len(cells))]
        count += len(cells)
        # row_data = [cell.text.strip() for cell in cells]
        
        # Check for rowspan and colspan attributes
        # r_count = {}
        # c_count = {}
        # for i in len(row_data):

        
        table_data.append(row_data)
    # print(cellss)
    merge= []
    for i, cells in enumerate(cellss):
        for j, cell in enumerate(cells):
            rowspan = int(cell.get('rowspan', 1))
            colspan = int(cell.get('colspan', 1))
            merge_row = rowspan > 1
            merge_col = colspan > 1
            # print(rowspan, colspan)
            # print(rowspan>1, colspan > 1)
            # merge_mix = rowspan > 1 and colspan > 1
            if merge_row:
                # print(cells)

                # print(f"check rows {i}, {rowspan}: {table_data[i+1:i+rowspan][:]}")
                if j < 3:
                    for row in table_data[i+1:i+rowspan]:
                        row.insert(j, [])
                    merge.append([i+1,j+1,i+rowspan,j+1])
                else:
                    for row in table_data[i+1:i+rowspan]:
                        row.insert(j, table_data[i][j])
                    
            if merge_col:
                # print("check col")
                for k in range(1, colspan):
                    table_data[i].insert(j+k, [])   
                merge.append([i+1,j+1,i+1,j+colspan])

        #     if merge_col:
        #         pass
        #     elif merge_row:
        #         pass
        #     elif merge_mix:

            # print(rowspan, colspan)
            # if rowspan > 1 or colspan > 1:
            #     # For cells with rowspan or colspan, repeat their content accordingly
            #     row_data.extend([cell.text.strip()] * (colspan - 1))
        
        # Append the row data to the table data list
        # while len(row_data) < 5:
        #     row_data.append('2')
        # print(row_data)
        list_a += table_data[i]
        print(table_data[i])
        # ws.append([str(c) for c in table_data[i]])
    
    out = match_boxes(list_a, rec_boxes, rec["text"])
    print(out)


    for i, row in enumerate(table_data):
        row = [out[5*i:5*i+5]]
        tmp = row
        row = []
        for item in tmp[0]:
            print(item)
            if len(item) > 1:
                item = reduce(lambda x, y: x + " " + y if len(x) > 0 and len(y) > 0 else x + y, item)
            else:
                item = item[0]
            row.append(item)
        rel = row[0] + " " + row[1]
        row[0] = (rel).replace(rel.split(' ')[0], "") if len(row[0]) > 0 else row[1]
        row.remove(row[1])
        ws.append(row)


    # for val in merge:
        # ws.merge_cells(start_row=val[0], start_column=val[1], end_row=val[2], end_column=val[3])
    # Create a DataFrame from the table data
    # df = pd.DataFrame(table_data)

    # # Remove empty columns from the DataFrame
    # df = df.loc[:, ~df.columns.duplicated()]
    # print(df.head(10))
    # # Create an Excel writer and write the DataFrame to it

    # for row in dataframe_to_rows(df, index=False, header=False):

    # # Perform cell merging in the Excel file based on merged_cells property
    # for cell_range in ws.merged_cells.ranges:
    #     start_row, start_col, end_row, end_col = cell_range.bounds
    #     if end_row - start_row > 0:
    #         print("merge here")
    #         ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=end_row, end_column=end_col)

    # Extend the cell to fit the content
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

    # Save the Excel file
    wb.save(excel_file)

    print(f"DataFrame saved to '{excel_file}' with cell merging.")

    # print("check", table_data, rec_boxes)


def calculate_center(box):
    x1, y1, x2, y2 = box
    x_center = (x1 + x2) // 2
    y_center = (y1 + y2) // 2
    return x_center, y_center

def match_text_with_cells(boxes_A, boxes_B, texts_B):
    matched_texts = []

    for box_A in boxes_A:
        xA, yA, xB, yB = box_A
        x_center_A, y_center_A = calculate_center(box_A)
        min_distance = float('inf')
        matched_text = None

        for i, box_B in enumerate(boxes_B):
            xC, yC, xD, yD = box_B
            x_center_B, y_center_B = calculate_center(box_B)

            # Calculate the distance between the centers of the two boxes
            distance = ((x_center_B - x_center_A) ** 2 + (y_center_B - y_center_A) ** 2) ** 0.5

            # Check if the box in list B is on the same row as the box in list A and closer in both x and y positions
            if yA <= y_center_B <= yB and distance < min_distance:
                min_distance = distance
                matched_text = texts_B[i]
                print(min_distance)

        matched_texts.append(matched_text)

    return matched_texts

def calculate_intersection(box1, box2):
    # Function to calculate the intersection area of two boxes
    # Assuming boxes are in the format [x1, y1, x2, y2]
    x_overlap = max(0, min(box1[2], box2[2]) - max(box1[0], box2[0]))
    y_overlap = max(0, min(box1[3], box2[3]) - max(box1[1], box2[1]))
    return x_overlap * y_overlap

def calculate_percentage_overlap(box1, box2):
    # Function to calculate the percentage overlap of two boxes
    intersection_area = calculate_intersection(box1, box2)
    box1_area = (box1[2] - box1[0]) * (box1[3] - box1[1])
    return intersection_area / box1_area

def match_boxes_kd_tree(list_a, list_b, distance_threshold=10):
    # Identify missing cells in list A
    missing_cells = [i for i, cell in enumerate(list_a) if not cell]

    # Remove missing cells from list A
    list_a_without_missing = [cell for cell in list_a if cell]

    # Convert box coordinates to 2D points for KDTree
    list_a_points = np.array([[(box[0] + box[2]) / 2, (box[1] + box[3]) / 2] for box in list_a_without_missing])
    list_b_points = np.array([[(box[0] + box[2]) / 2, (box[1] + box[3]) / 2] for box in list_b])

    # Construct KD tree for list B
    b_kd_tree = KDTree(list_b_points)

    # Perform the matching for non-empty cells in list A
    matched_indices = b_kd_tree.query(list_a_points)[1]

    # Initialize results list
    results = [None] * len(list_a)

    # Fill in the matched cells in list A with the corresponding text from list B
    for i, matched_index in enumerate(matched_indices):
        results[i] = list_b[matched_index]

    # Fill in missing cells with None or a default value
    for missing_index in missing_cells:
        if results[missing_index] is None:
            results[missing_index] = None  # Or use a default value

    return results

# from scipy.spatial import KDTree

# def calculate_intersection(box1, box2):
#     # Function to calculate the intersection area of two boxes
#     # Assuming boxes are in the format [x1, y1, x2, y2]
#     x_overlap = max(0, min(box1[2], box2[2]) - max(box1[0], box2[0]))
#     y_overlap = max(0, min(box1[3], box2[3]) - max(box1[1], box2[1]))
#     return x_overlap * y_overlap

# def calculate_percentage_overlap(box1, box2):
#     # Function to calculate the percentage overlap of two boxes
#     intersection_area = calculate_intersection(box1, box2)
#     box1_area = (box1[2] - box1[0]) * (box1[3] - box1[1])
#     return intersection_area / box1_area

# def match_boxes_kd_tree(list_a, list_b):
#     # Identify missing cells in list A
#     missing_cells = [i for i, cell in enumerate(list_a) if not cell]

#     # Remove missing cells from list A
#     list_a_without_missing = [cell for cell in list_a if cell]

#     # Construct KD tree for list B
#     b_kd_tree = KDTree(list_b)

#     # Perform the matching for non-empty cells in list A
#     matched_indices = b_kd_tree.query(list_a_without_missing)[1]

#     # Initialize results list
#     results = [None] * len(list_a)

#     # Fill in the matched cells in list A
#     for i, matched_index in enumerate(matched_indices):
#         results[i] = list_b[matched_index]

#     # Fill in missing cells with None or a default value
#     for missing_index in missing_cells:
#         results.insert(missing_index, None)  # Or use a default value

#     return results


if __name__== "__main__":
    to_excel()
    # # print("main")

    # # # Example usage:
    # # boxes_A = [[723, 1, 824, 74],[], [867, 3, 1020, 74], [1091, 1, 1237, 34]]
    # # boxes_B = [[727, 0, 827, 31], [890, 0, 1021, 31], [1095, 0, 1241, 31]]
    # # texts_B = ["Text 1", "Text 2", "Text 3"]

    # # matched_texts = match_text_with_cells(boxes_A, boxes_B, texts_B)
    # # print(matched_texts)

    # a = open("/home/xuan/Project/OCR/sample/result/pred/res_1_TSR.txt", 'r')
    # b = open("/home/cxuan/Project/OCR/sample/result/text/1/res.json", "r")

    # res_a = json.load(a)
    # res_b = json.load(b)

    # # Example usage
    # # list_a = [[1, 2, 3, 4], [], [5, 6, 7, 8], [9, 10, 11, 12], [], [13, 14, 15, 16]]
    # # list_b = [[17, 18, 19, 20], [21, 22, 23, 24], [25, 26, 27, 28], [29, 30, 31, 32], [33, 34, 35, 36]]
    # list_a = res_a["content_ann"]["bboxes"]
    # ratio = res_a["ratio"]
    # tmp = list_a
    # list_a = []
    # for box in tmp:
    #     box = np.asarray(box).astype(np.float64)
    #     box *= ratio
    #     box = list(box.astype(np.int64))
    #     list_a.append(box)
    
    # list_b = res_b["bboxes"]

    # matches = match_boxes_kd_tree(list_a, list_b)
    # print(matches)